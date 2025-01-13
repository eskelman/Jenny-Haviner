# %%
print("Doing my thing...\n")

import pandas as pd
import os
import glob
from datetime import datetime
import re
import shutil
import xlsxwriter
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Side, Border
import calendar
import numpy as np

status = 'open' # Just means you can close the tab and assess any errors
while status == 'open':
    
# %%
 
    # Gather specified columns
    def gather_calendar():
        
        # Set location to find Data, this must be in the same file path as the exe
        input_source = "Data\input\Gigabook Calendar"
        list_of_files = glob.glob(f'{input_source}/*.csv') # glob is an efficient file finding list
        latest_input_file = max(list_of_files, key=os.path.getctime) # Always gets the most recent file

        # Load data in to a DataFrame
        df=pd.read_csv(latest_input_file,header=0)
        condensed_table = df.loc[:,["Instructor Name", "Description", "Date/Time", "Duration", "Provider Note"]]
        num_lessons = len(condensed_table)
        return condensed_table, num_lessons

    pay_data, num_lessons = gather_calendar()
    

    # %%
    
    # Get time and date for file naming and selecting the correct pay range sheet
    date = pay_data["Date/Time"]
    def extract_date(text):
        result = re.search('([^*]+) @', text).group(1)
        return result
    dates = date.apply(lambda x: extract_date(x))
    dates = pd.to_datetime(dates)

    # %%
    
    # Find commission COULD SAVE IN INTERIM LOCATION
    def id_commission(pay_data):
        #pattern = r'\bcom{1,2}ission - \$(\d*.\d*)?\b' # this regex requires the commission to be entered "commission - ${digits}". The term is case insensitive and the number can have a either have a "." or not
        pattern = r'\$(\d*.\d*)?\b' # Retrieves any term after the $, which leaves it open to error
        out = pay_data['Provider Note'].str.extract(pattern, flags=re.IGNORECASE)
        
        try:
            pay_data.drop('Commission',axis = 1) # on the event there is already a Commission, it deletes and replaces it 
        except:
            pay_data.insert(pay_data.columns.get_loc('Provider Note')+1, 'Commission', out[0])
            
        return pay_data

    pay_data_with_commission = id_commission(pay_data)

    # %%
    
    # Work out the pay rate year. This is why the naming convention for the Pay Rates sheet needs to be standardised
    def pay_rate_year(dates):
        year = int(dates[1].year) - 2000 # Clumsy, I know. If you haven't got a better system by the next century - that's on you!
        month = int(dates[1].month) 
        if month <= 6:
            pay_period = f"{year-1}-{year}"
        if month > 6:
            pay_period = f"{year}-{year+1}"
        return pay_period
    pay_period = pay_rate_year(dates)

    # %%
    
    # Get the rates of pay to refer to
    def gather_pay_rates(pay_period):
        input_source = "Data\input\Pay Rates"
        list_of_files = glob.glob(f'{input_source}/*.xlsx') # glob is an efficient file finding lib
    
        latest_input_file = max(list_of_files, key=os.path.getctime) # Always gets the most recent file
     
        concat_rates=pd.read_excel(latest_input_file,f"{pay_period} Rates", header=None) # So long as the Pay Rates sheet uses the same naming conversion for sheets, you should be able to have multiple years in the same workbook
        return concat_rates

    pay_rates_sheet = gather_pay_rates(pay_period)

    # %%

    # Identify the SKI rate section of the sheet
    def gather_ski_rates(pay_rates_sheet):
        ski_index = pay_rates_sheet[pay_rates_sheet.iloc[:,0].str.lower().str.contains('ski',na=False)].index # looks for the work SKI then takes the table below it, until an empty cell

        if ski_index[0] < len(pay_rates_sheet):
            extracted_rows = pay_rates_sheet.iloc[ski_index[0] + 1:]
            
            ski_rates = pd.DataFrame()
            
            # Iterate through rows until column 1 becomes empty
            for index, row in extracted_rows.iterrows():
                if pd.isnull(row.iloc[0]):  # Check if the first column is empty
                    break
                ski_rates = pd.concat([ski_rates, row.to_frame().transpose()], ignore_index=True)
        
        return ski_rates

    ski_rates = gather_ski_rates(pay_rates_sheet)
    ski_rates.columns = ski_rates.iloc[0] #Would be nice to get these in the function
    ski_rates = ski_rates.drop(ski_rates.index[0])

    # %%
    
    # Identify the BOARD rate section of the sheet
    def gather_board_rates(pay_rates_sheet):
        board_index = pay_rates_sheet[pay_rates_sheet.iloc[:,0].str.lower().str.contains('board',na=False)].index

        if board_index[0] < len(pay_rates_sheet):
            extracted_rows = pay_rates_sheet.iloc[board_index[0] + 1:]

            
            board_rates = pd.DataFrame()
            # Iterate through rows until column 1 becomes empty
            for index, row in extracted_rows.iterrows():
                if pd.isnull(row.iloc[0]):  # Check if the first column is empty
                    break
                board_rates = pd.concat([board_rates, row.to_frame().transpose()], ignore_index=True)
        return board_rates

    board_rates = gather_board_rates(pay_rates_sheet)
    board_rates.columns = board_rates.iloc[0] #Would be nice to get these in the function
    board_rates = board_rates.drop(board_rates.index[0])

    # %%

    def calc_dual_rates(ski_rates, board_rates):
        # Find dual instructors and their rates
        dual_rates = pd.DataFrame(columns=['Lesson Type'])
        
        for instructor in board_rates.columns[1:]:
            if instructor in ski_rates.columns:
                concat_rates = pd.concat([board_rates[['Lesson Type', instructor]], ski_rates[['Lesson Type', instructor]]])
                
                # Extract the Private Lessons for ski and board to add at the end
                privates = concat_rates[concat_rates['Lesson Type'].str.contains(r'Private \d', regex=True)]
                
                # Extract the common lesson types without (Board) or (Ski)
                concat_rates['Lesson Type'] = concat_rates['Lesson Type'].str.extract(r'(.+?)\s*\(.+?\)')

                # Initialize a new dataframe for dual rates
                dual_rates_instructor = pd.DataFrame(columns=['Lesson Type', instructor])

                # Iterate through unique lesson types
                for lesson_type in concat_rates['Lesson Type'].unique():
                    
                    # Get rows for the current lesson type
                    rows = concat_rates[concat_rates['Lesson Type'] == lesson_type]

                    # Find the maximum rate
                    max_rate = rows[instructor].max()

                    # Get the row with the maximum rate
                    max_row = rows[rows[instructor] == max_rate].iloc[0].copy()

                    # Rename the lesson type to "Lesson Type (Dual)"
                    max_row['Lesson Type'] = f"{lesson_type} (Dual)"

                    # Add the row to the dual_rates_instructor dataframe
                    dual_rates_instructor = pd.concat([dual_rates_instructor, max_row.to_frame().transpose()])

                # Remove duplicates from dual_rates_instructor
                dual_rates_instructor = dual_rates_instructor.drop_duplicates(subset='Lesson Type', keep='first')
                dual_rates_instructor = pd.concat([dual_rates_instructor, privates])
                
                # Concatenate the results for each instructor into the final dual_rates dataframe
                dual_rates = dual_rates.merge(dual_rates_instructor, on='Lesson Type', how='right')

        return dual_rates

    dual_rates = calc_dual_rates(ski_rates, board_rates)

    # %%
    
    # Group all lessons by Instructor
    def group_lessons(pay_data_with_commission):
        #instructor_list = pay_data_with_commission["Instructor Name"].unique()

        # Grouping by 'Instructor Name'
        grouped = pay_data_with_commission.groupby("Instructor Name")

        # Creating individual DataFrames for each group
        individual_dfs = {name: group for name, group in grouped}
        instructor_dict = {}

        # Create dictionary for each instructor's lessons
        for instructor_name, individual_df in individual_dfs.items():
            instructor_dict[instructor_name] = individual_df
        return instructor_dict
    
    instructor_dict = group_lessons(pay_data_with_commission)

    # %%
    ################# Functions to call on in the iterative process #########################
    
    # Identify the instructor's discipline
    def instructor_discipline(instructor,ski_rates,board_rates):

        if instructor in board_rates.columns and instructor in ski_rates.columns:
            itype = "dual"
        elif instructor in ski_rates.columns:
            itype = "ski"
        elif instructor in board_rates.columns:
            itype = "board"
        
        else:
            itype = 'unknown' 
            # Give indication as to which is the unknown instructor (usually a wrong name spelling)
            print(f"{instructor}: unknown\n- Do the instructor spellings on the Pay Rates match the Calendar\n- Is the instructor on the Pay Rates sheet")
        return itype


    # %%
    
    # Formatting rough instructor table into time sheet
    def format_calendar(instructor_calendar, discipline):
        
        # Delete provider note column
        instructor_calendar = instructor_calendar.drop(["Provider Note"], axis=1)
        # Remove 'hr' from duration and convert to int
        instructor_calendar["Duration"] = instructor_calendar["Duration"].apply(lambda x : int(x[:-3]))
        # Remove the time and year from Date/Time
        
        def extract_date(text):
            #result = re.search("([^,]+)", text)
            result = re.search("([a-zA-Z]{3}) 0?(\d{1,2})", text)
            return f"{result.group(1)} {result.group(2)}"
        
        instructor_calendar.sort_values(by=['Date/Time'], inplace = True)
        instructor_calendar ["Date/Time"] = instructor_calendar["Date/Time"].apply(lambda x: extract_date(x))

        # Change the lesson type to match the format in the timesheet
        def match_lesson_type(text,discipline,name):
            if discipline == 'dual':
                
                p6_ski = ["Full-Day Ski Private Lesson"]
                p3_ski = ["Half-Day Ski Private Lesson"]
                p6_board = ["Full-Day Snowboarding Private Lesson"]
                p3_board = ["Half-Day Snowboarding Private Lesson"]
                p6_dual = ["Guided Adventure 6 HR"]
                p3_dual = ["Guided Adventure 3 HR"]
                NS = ["NonStop"]
                training = ["Training"]
                trainer = ["Trainer"]
                SB = ["Stand By"]
                groups_dual = ["School Groups"]
            
            
                if any(x in text for x in p6_dual):
                        result = "Private 6 (Dual)"
                elif any(x in text for x in p3_dual):
                        result = "Private 3 (Dual)"
                elif any(text in x for x in NS):
                        result = "NonStop (Dual)"
                elif any(x in text for x in SB):
                        result = "Stand by (Dual)"
                elif any(text in x for x in training):
                        result = "Training (Dual)"
                elif any(text in x for x in trainer):
                        result = "Trainer (Dual)"
                elif any(text in x for x in groups_dual):
                    result = "Groups (Dual)"
                elif any(x in text for x in p6_ski):
                    result = "Private 6 (Ski)"
                elif any(x in text for x in p3_ski):
                    result = "Private 3 (Ski)"
                elif any(x in text for x in p6_board):
                    result = "Private 6 (Board)"
                elif any(x in text for x in p3_board):
                    result = "Private 3 (Board)"
                else:
                    print(f"Cant find a match, description needs editing\n- {name}\n- {text}")
                    result = "Unmatched"
        
            if discipline == 'ski':
                p6_ski = ["Full-Day Ski Private Lesson","Guided Adventure 6 HR"]
                p3_ski = ["Half-Day Ski Private Lesson", "Guided Adventure 3 HR"]
                NS = ["NonStop"]
                training = ["Training 3 Hrs", "Training 6 Hrs"]
                trainer = ["Trainer"]
                SB = ["Stand By"]
                groups_ski = ["School Groups"]
                
                if any(x in text for x in p6_ski):
                    result = "Private 6 (Ski)"
                elif any(x in text for x in p3_ski):
                    result = "Private 3 (Ski)"
                elif any(text in x for x in NS):
                    result = "NonStop (Ski)"
                elif any(x in text for x in SB):
                        result = "Stand by (Ski)"
                elif any(text in x for x in training):
                    result = "Training (Ski)"
                elif any(text in x for x in trainer):
                    result = "Trainer (Ski)"
                elif any(text in x for x in groups_ski):
                    result = "Groups (Ski)"
                else:
                    print(f"Cant find a match, description needs editing\n- {name}\n- {text}")
                    result = "Unmatched"
                
            if discipline == 'board':
                p6_board = ["Full-Day Snowboarding Private Lesson","Guided Adventure 6 HR"]
                p3_board = ["Half-Day Snowboarding Private Lesson", "Guided Adventure 3 HR"]
                NS = ["NonStop"]
                training = ["Training 3 Hrs", "Training 6 Hrs"]
                trainer = ["Trainer"]
                SB = ["Stand By"]
                groups_board = ["School Groups"]
                    
                if any(x in text for x in p6_board):
                    result = "Private 6 (Board)"
                elif any(x in text for x in p3_board):
                    result = "Private 3 (Board)"
                elif any(text in x for x in NS):
                    result = "NonStop (Board)"
                elif any(x in text for x in SB):
                        result = "Stand by (Board)"
                elif any(text in x for x in training):
                    result = "Training (Board)"
                elif any(text in x for x in trainer):
                    result = "Trainer (Board)"
                elif any(text in x for x in groups_board):
                    result = "Groups (Board)"
                else:
                    print(f"Cant find a match, description needs editing\n- {name}\n- {text}")
                    result = "Unmatched"
                
            return result
            
            
        instructor_calendar["Description"] = instructor_calendar["Description"].apply(lambda x: match_lesson_type(x,discipline,instructor_calendar['Instructor Name'].values[0]))
    
        # Combine the hours of any duplicates - i.e. two of the same lesson type on the same day
        instructor_calendar['Agg Duration'] = instructor_calendar.groupby(['Instructor Name', 'Description', 'Date/Time'])['Duration'].transform('sum')
        instructor_calendar['Agg Commission'] = instructor_calendar.groupby(['Instructor Name', 'Description', 'Date/Time'])['Commission'].transform('sum')

        # Drop the original 'Duration' and 'Commission' columns
        instructor_calendar.drop(['Duration', 'Commission'], axis=1, inplace=True)

        # Drop duplicates to keep only unique rows, rename columns back to "Duration" and "Commission"
        instructor_calendar.drop_duplicates(inplace=True)
        instructor_calendar = instructor_calendar.rename(columns={"Agg Duration": "Duration", "Agg Commission": "Commission"})
        
        return instructor_calendar


    # %%
    def format_rates(instructor, ski_rates, board_rates, dual_rates):
        # Group all rates into a df
        ins_ski_rates = pd.DataFrame()
        ins_board_rates = pd.DataFrame()
        ins_dual_rates = pd.DataFrame()

        if instructor in board_rates.columns and instructor in ski_rates.columns:
            ins_dual_rates = dual_rates[["Lesson Type", instructor]]

        elif instructor in ski_rates.columns:
            ins_ski_rates = ski_rates[["Lesson Type", instructor]]

        elif instructor in board_rates.columns:
            ins_board_rates = board_rates[["Lesson Type", instructor]]

        instructor_rates = pd.concat([ins_ski_rates, ins_board_rates, ins_dual_rates], ignore_index=True).rename(columns={instructor:"Rates"})
        instructor_rates.loc[len(instructor_rates.index)] = ["Commission"]+[None] * (len(instructor_rates.columns)-1)
        
        # Insert the dates in columns between the lesson type and rates
        sorted_dates = dates.sort_values(ascending=False)
        date_list = [f"{calendar.month_abbr[x.month]} {x.day}" for x in sorted_dates.unique()]
        instructor_timesheet = instructor_rates
        for column in date_list:
            instructor_timesheet.insert(1, column, None)
        
        return instructor_timesheet

    #format_rates("Carlo Riveroll", ski_rates, board_rates, dual_rates)

    # %%
    def create_timesheet(formatted_calendar, formatted_rates, discipline):
        # set the suffix used from discipline
        if discipline == 'ski':
            suffix = '(Ski)'
        if discipline == 'board':
            suffix = '(Board)'
        if discipline == 'dual':
            suffix = '(Dual)'
            
        # list each calendar row [lesson type, date, duration, commission]
        for index, row in formatted_calendar[["Instructor Name","Description","Date/Time","Duration","Commission"]].iterrows():
            lt = row["Description"]
            dt = row["Date/Time"]
            dur = row["Duration"]
            com = row["Commission"]
            nm = row["Instructor Name"]
            
            # Find the corresponding row index for the lesson type and commission in instructor calendar
            try:
                lt_index = formatted_rates.index[formatted_rates['Lesson Type'] == f"{lt}"].tolist()[0]
            except:
                print(f"Cannot locate the lesson combination:\nInstructor: {nm} \nLesson Type: {lt}\nDuration: {dur}\nDate: {dt}\n")
                lt_index = "na"

            com_index = formatted_rates.index[formatted_rates["Lesson Type"] == "Commission"]

            # Update the final timesheet with the duration value at corresponding date and lesson type
            if lt_index != 'na':
                formatted_rates.loc[lt_index, dt] = dur
            if com != 0:
                formatted_rates.loc[com_index, dt] = com
        
        # Create a new column "Total" in instructor_timesheet
        formatted_rates['Total'] = 0.0
        formatted_rates.insert(0,"Instructor Name",formatted_calendar["Instructor Name"].values[0],True)

        # Iterate through rows
        for index, row in formatted_rates.iterrows():
            # Protocol for lesson type rows
            if index != formatted_rates.index[-1]:
                total = 0.0

                # Iterate through date columns in instructor_timesheet and calculate the sum for each date
                for column in formatted_rates.columns[2:-2]:  # Exclude the first column and the last two columns ('Rates' and 'Total')
                    # Replace None or NaN with 0 before performing the multiplication
                    row_sum = 0
                    if pd.notna(row[column]):
                        row_sum += pd.to_numeric(row[column])

                    # Multiply the sum by the corresponding rate
                    rate = formatted_rates.at[index, 'Rates']
                    if pd.notna(rate):
                        total += row_sum * rate

                # Update the sum duration * rate in the total column
                formatted_rates.at[index, 'Total'] = total
            
            # Protocol for commission row    
            if index == formatted_rates.index[-1]:
                commission_sum = 0
                for column in formatted_rates.columns[2:-2]:
                    if pd.notna(row[column]):
                        commission_sum += pd.to_numeric(row[column])
                
                # Update the sum commission in the total column
                formatted_rates.at[index, 'Total'] = commission_sum

        # Add a new cell under the total column to get the grand total
        new_row = pd.Series(dtype='float64')
        formatted_rates.loc[len(formatted_rates)] = new_row
        sum_total = formatted_rates.iloc[:-1, -1].sum()
        formatted_rates.at[len(formatted_rates)-1, formatted_rates.columns[-1]] = sum_total
        return formatted_rates


    # %%
    def add_to_timesheet(collective_timesheet, instructor_timesheet):
        # add blank row to instructor timesheet
        blank_row = pd.Series(dtype='float64')
        instructor_timesheet.loc[len(instructor_timesheet)] = blank_row
        instructor_timesheet.loc[len(instructor_timesheet)] = blank_row
        
        # add instructor timesheet to complete timesheet
        complete_timesheet = pd.concat([collective_timesheet, instructor_timesheet], ignore_index=True)
        return complete_timesheet


    # %%
    def export_timesheet(collective_timesheet,dates):
        max_date = f"{calendar.month_abbr[max(dates).month]} {max(dates).day}"
        min_date = f"{calendar.month_abbr[min(dates).month]} {min(dates).day}"
        date_range = f"{min_date} {min(dates).year} - {max_date} {max(dates).year}"
        how_many_dates = len(dates.unique())
        
        # Load the existing Excel file
        path = './Data/interim/timesheet_template.xlsx'
        new_name = f'./Data/output/timesheet_{min_date}_{max_date}.xlsx'
        shutil.copy(path, new_name)
        
        # Get the ExcelWriter object

        with pd.ExcelWriter(new_name, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
            workbook = writer.book
            worksheet = writer.sheets['Timesheet']
            
            # Format for dataframe input (fonts, bold ...)
            dotum_format = Font(bold=True, color="000000", name="Dotum")
            red_format = PatternFill(start_color="FF0000", end_color="FF0000", fill_type = "solid")
            
            # Input the date at the top
            worksheet['A2'].font = dotum_format
            worksheet['A2'] = date_range

            # Write the DataFrame to the Excel file starting from row 10
            collective_timesheet.to_excel(writer, sheet_name='Timesheet', startrow=9, header=True, index=False)
            
            # Apply font to the timesheet area of the workbook
            start_row = 9
            index = start_row
            for row in worksheet.iter_rows(min_row=start_row, min_col=0, max_row=worksheet.max_row, max_col=worksheet.max_column):
    
                for cell in row:
                    cell.font = dotum_format
                
                # Highlight Rates that need to be checked
                highlight = ['Groups (Dual)', 'Training (Dual)', 'Trainer (Dual)','NonStop (Dual)']
                  
                if row[1].value in highlight:
                    row[how_many_dates+2].fill = red_format
                    
                if row[1].value == 'Commission':
                    for col_num in range(1, how_many_dates + 5):  # Iterate over columns
                        cell = worksheet.cell(row=index+1, column=col_num)
                        cell.border = Border(left=Side(style='none'), 
                                            right=Side(style='none'), 
                                            top=Side(style='thin'), 
                                            bottom=Side(style='thin'))
                
                
                index += 1
                    


    # %%
    # This is the point in which we iterate through instructors in instructor dict
    collective_timesheet = pd.DataFrame()
    lesson_count = 0
    
    for instructor in list(instructor_dict)[:]:
        instructor_calendar = instructor_dict[instructor]

        # ID the intructor's discipline
        discipline = instructor_discipline(instructor,ski_rates,board_rates)
        if discipline != 'unknown':
            print(f'{instructor}: {discipline}')
            
            # Format the calendar and rates for further use
            formatted_calendar = format_calendar(instructor_calendar, discipline)
            formatted_rates = format_rates(instructor, ski_rates, board_rates, dual_rates)
            lesson_count += len(formatted_calendar)

            # Group everything together into a timesheet
            formatted_timesheet = create_timesheet(formatted_calendar, formatted_rates, discipline)
            
            # Add individual instructor timesheet to a collective timesheet
            collective_timesheet = add_to_timesheet(collective_timesheet, formatted_timesheet)
            
            
        else:
            print(f'{instructor} omitted due to lack of data\n')
            continue
        
    # Export collective timesheet to csv
    export_timesheet(collective_timesheet,dates)
        
    # %%
    ############## Run Checks ###############
    
    if num_lessons == lesson_count:
        print(f"\nNumber of lessons check - Pass\n")
    else:
        print(f"\nNumber of lessons check - Fail\n")


    # %%
    print("\nFinished!\nYou can access the Timesheet now.\n")
    status = input('Type \'c\' to shut down the script')